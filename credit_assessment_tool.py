import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import os
from fpdf import FPDF
from io import BytesIO
from openpyxl import Workbook

st.title("📝 Credit Assessment Tool")

if 'page' not in st.session_state:
    st.session_state.page = 1

if 'retailer_data' not in st.session_state:
    st.session_state.retailer_data = {}

##############################################
# PAGE 1: Full Financial Data Entry
##############################################
if st.session_state.page == 1:
    st.header("1️⃣ Enter Full Financial Information")

    # Retailer Basic Info
    st.subheader("Retailer Details")
    state = st.text_input("State")
    district = st.text_input("District")
    retailer_name = st.text_input("Retailer Name")
    sap_code = st.text_input("SAP Code")
    file_number = st.text_input("File Number")
    dealing_period = st.selectbox("Dealing Period",
                                  ["More Than 5 years Active", "From 3 to 5 Years Active", "From 1 to 3 years Active",
                                   "Less than 1 year Active", "New Dealer"])

    # Balance Sheet Inputs
    st.subheader("Balance Sheet (Amounts in Rs.)")
    col1, col2 = st.columns(2)
    with col1:
        owners_capital = st.number_input("Owners' Capital Account")
        reserves = st.number_input("Reserves and Surplus")
        long_term_borrowings = st.number_input("Long-term Borrowings")
        deferred_tax_liabilities = st.number_input("Deferred Tax Liabilities (Net)")
        other_long_term_liabilities = st.number_input("Other Long-term Liabilities")
        short_term_borrowings = st.number_input("Short-term Borrowings")
        trade_payables = st.number_input("Trade Payables")
        other_current_liabilities = st.number_input("Other Current Liabilities")
        short_term_provisions = st.number_input("Short-term Provisions")
    with col2:
        prop_plant_eqp = st.number_input("Property, Plant and Equipment")
        intangible_assets = st.number_input("Intangible Assets")
        capital_work_progress = st.number_input("Capital Work in Progress")
        non_current_investments = st.number_input("Non-current Investments")
        inventories = st.number_input("Inventories")
        trade_receivables = st.number_input("Trade Receivables")
        cash_bank = st.number_input("Cash and Bank Balances")
        other_current_assets = st.number_input("Other Current Assets")

    # P&L Inputs
    st.subheader("Profit & Loss (Amounts in Rs.)")
    revenue = st.number_input("Revenue from Operations")
    other_income = st.number_input("Other Income")
    cost_goods_sold = st.number_input("Cost of Goods Sold")
    finance_costs = st.number_input("Finance Costs")
    depreciation = st.number_input("Depreciation and Amortization")
    other_expenses = st.number_input("Other Expenses")

    # Uploads
    st.subheader("Upload Financial Statements & Other Docs (PDF)")
    uploaded_file = st.file_uploader("Upload a PDF file")

    col_next, col_exit = st.columns([1, 1])
    if col_next.button("Next ➡️"):
        st.session_state.retailer_data = {
            'State': state,
            'District': district,
            'Retailer Name': retailer_name,
            'SAP Code': sap_code,
            'File Number': file_number,
            'Dealing Period': dealing_period,
            'Owners Capital': owners_capital,
            'Reserves': reserves,
            'Long Term Borrowings': long_term_borrowings,
            'Deferred Tax Liabilities': deferred_tax_liabilities,
            'Other Long Term Liabilities': other_long_term_liabilities,
            'Short Term Borrowings': short_term_borrowings,
            'Trade Payables': trade_payables,
            'Other Current Liabilities': other_current_liabilities,
            'Short Term Provisions': short_term_provisions,
            'Property Plant Equipment': prop_plant_eqp,
            'Intangible Assets': intangible_assets,
            'Capital Work Progress': capital_work_progress,
            'Non Current Investments': non_current_investments,
            'Inventories': inventories,
            'Trade Receivables': trade_receivables,
            'Cash Bank': cash_bank,
            'Other Current Assets': other_current_assets,
            'Revenue': revenue,
            'Other Income': other_income,
            'Cost of Goods Sold': cost_goods_sold,
            'Finance Costs': finance_costs,
            'Depreciation': depreciation,
            'Other Expenses': other_expenses,
            'Uploaded File': uploaded_file
        }
        folder_name = f"Retailers/{state.replace(' ', '_')}/{district.replace(' ', '_')}/{retailer_name.replace(' ', '_')}"
        os.makedirs(folder_name, exist_ok=True)
        if uploaded_file:
            with open(os.path.join(folder_name, uploaded_file.name), "wb") as f:
                f.write(uploaded_file.getbuffer())
        st.session_state.page = 2
        st.rerun()

##############################################
# PAGE 2: Ratio Calculation & Scoring
##############################################
if st.session_state.page == 2:
    st.header("2️⃣ Ratios & Financial Scoring")

    data = st.session_state.retailer_data

    # ---- Calculations ----
    TOL = (data['Long Term Borrowings'] + data['Short Term Borrowings'] +
           data['Other Current Liabilities'] + data['Trade Payables'])
    TNW = data['Owners Capital'] + data['Reserves']
    tol_tnw = TOL / TNW if TNW != 0 else 0

    current_assets = (data['Inventories'] + data['Trade Receivables'] +
                      data['Cash Bank'] + data['Other Current Assets'])
    current_liabilities = (data['Short Term Borrowings'] + data['Trade Payables'] +
                           data['Other Current Liabilities'] + data['Short Term Provisions'])
    current_ratio = current_assets / current_liabilities if current_liabilities != 0 else 0

    total_income = data['Revenue'] + data['Other Income']
    pbdt = total_income - data['Cost of Goods Sold'] - data['Other Expenses']
    pbdt_int = pbdt / data['Finance Costs'] if data['Finance Costs'] != 0 else 0

    pat = total_income - (data['Cost of Goods Sold'] + data['Finance Costs'] +
                          data['Depreciation'] + data['Other Expenses'])
    net_cash_accrual = pat + data['Depreciation']
    total_debt = data['Long Term Borrowings'] + data['Short Term Borrowings']
    net_cash_td = (net_cash_accrual / total_debt) * 100 if total_debt != 0 else 0

    asset_turnover = (data['Revenue'] / (data['Inventories'] + data['Trade Receivables'])
                      if (data['Inventories'] + data['Trade Receivables']) != 0 else 0)

    # Display ratios
    ratios = {
        'TOL/TNW': round(tol_tnw, 2),
        'Current Ratio': round(current_ratio, 2),
        'PBDIT/Interest': round(pbdt_int, 2),
        'Net Cash Accruals/Total Debt (%)': round(net_cash_td, 2),
        'Asset Turnover': round(asset_turnover, 2)
    }
    st.subheader("🔢 Calculated Ratios")
    st.table(ratios)

    st.session_state.retailer_data['Ratios'] = ratios

    col_back, col_next = st.columns([1, 1])
    if col_back.button("⬅️ Back"):
        st.session_state.page = 1
        st.rerun()
    if col_next.button("Next ➡️"):
        st.session_state.page = 3
        st.rerun()

##############################################
# PAGE 3: Final Report Generation
##############################################
if st.session_state.page == 3:
    st.header("3️⃣ Generate PDF & Excel Report")

    data = st.session_state.retailer_data
    ratios = data['Ratios']

    st.subheader("Business & Managerial Inputs")
    business_score = st.slider("Business Score (0-25)", 0, 25, 15)
    managerial_score = st.slider("Managerial Score (0-10)", 0, 10, 5)
    quantity_increase = st.slider("Quantity Increase (0-10)", 0, 10, 5)
    analyst = st.text_input("Analysed By")
    remarks = st.text_area("Remarks")

    if st.button("Generate Reports ✅"):
        folder_name = f"Retailers/{data['State'].replace(' ', '_')}/{data['District'].replace(' ', '_')}/{data['Retailer Name'].replace(' ', '_')}"
        final_score = (ratios['TOL/TNW'] + ratios['Current Ratio'] + ratios['PBDIT/Interest'] +
                       ratios['Net Cash Accruals/Total Debt (%)'] + ratios['Asset Turnover']) + \
                      business_score + managerial_score + quantity_increase

        # PDF report
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Credit Assessment Report", ln=True, align="C")
        pdf.ln(10)
        pdf.multi_cell(0, 10, txt=f"State: {data['State']}\nDistrict: {data['District']}\nRetailer: {data['Retailer Name']}\nSAP Code: {data['SAP Code']}\n"
                                   f"File Number: {data['File Number']}\nDealing Period: {data['Dealing Period']}")
        pdf.ln(5)
        pdf.cell(200, 10, txt=f"Final Score: {round(final_score, 2)} | Analyst: {analyst}", ln=True)
        pdf.multi_cell(0, 10, txt=f"Remarks: {remarks}")
        pdf.ln(10)
        pdf.cell(200, 10, txt="Calculated Ratios:", ln=True)
        for k, v in ratios.items():
            pdf.cell(200, 10, txt=f"{k}: {v}", ln=True)
        pdf_file_path = f"{folder_name}/Final_Report.pdf"
        pdf.output(pdf_file_path)

        # Excel report
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Report"

        ws.append(["State", data['State']])
        ws.append(["District", data['District']])
        ws.append(["Retailer Name", data['Retailer Name']])
        ws.append(["SAP Code", data['SAP Code']])
        ws.append(["File Number", data['File Number']])
        ws.append([])
        ws.append(["Ratios"])
        for k, v in ratios.items():
            ws.append([k, v])

        excel_file_path = f"{folder_name}/Final_Report.xlsx"
        wb.save(excel_file_path)

        st.success(f"✅ PDF & Excel reports saved in: {folder_name}")
